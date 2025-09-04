import openpyxl
#22496
def remove_duplicates(file_path, sheet_name):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # 初始化存储唯一数据的集合
    # seen_signatures = set()
    count = 1
    # 从最后一行开始向前遍历，防止在删除行时出现索引问题
    for row in range(sheet.max_row, 1, -1):  # 跳过第一行标题
        cell = sheet.cell(row=row, column=1)
        value = cell.value
        
        if value and "条码"  and '算力云正向数据导出JSON中的字段为空，' and '指令类型' and '指令内容' in value :
            # 忽略具体物料编号部分, 只比较前缀
            # signature = "物料内存()数量"
            if count ==1:
                count= count+1
                continue
            else: 
                sheet.delete_rows(row)
            # else:
            #     seen_signatures.add(signature)

    # 保存修改后的文件
    workbook.save('test02.xlsx')
    print("数据处理完成，文件已保存。")

# 调用函数
remove_duplicates('test02.xlsx', 'Sheet2')
